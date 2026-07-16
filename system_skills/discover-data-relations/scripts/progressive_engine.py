#!/usr/bin/env python3
"""Progressive relationship discovery without full-value indexing.

The engine inventories structure, extracts bounded seeds from small/high-signal files,
probes candidate columns in larger tables, and persists only bounded profiles and
confirmed evidence. Raw source content is never written to analysis artifacts.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import itertools
import json
import os
import re
import sqlite3
import subprocess
import sys
import tempfile
import time
import unicodedata
import zipfile
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator, Sequence
from xml.etree import ElementTree


SCHEMA_VERSION = 3
TABULAR_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".jsonl", ".ndjson", ".parquet", ".sqlite", ".sqlite3", ".db"}
TEXT_EXTENSIONS = {".txt", ".md", ".markdown", ".log", ".sql", ".json", ".xml", ".html", ".htm", ".yaml", ".yml", ".ini", ".cfg", ".conf"}
DOCUMENT_EXTENSIONS = TEXT_EXTENSIONS | {".pdf", ".docx", ".pptx"}
OCR_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff", ".webp"}
JOIN_KINDS = {"id", "code", "email", "phone", "url", "reference", "name"}
STRONG_KINDS = {"id", "code", "email", "phone", "url", "uuid", "labeled_value"}
ROLE_TERMS = {
    "rule", "rules", "policy", "policies", "config", "configuration", "reference", "lookup",
    "dictionary", "mapping", "template", "sample", "example", "result", "output", "spec",
    "规则", "政策", "配置", "参考", "字典", "映射", "模板", "样例", "示例", "结果", "输出", "规范",
}
MAX_EVIDENCE = 12


@dataclass
class ColumnMeta:
    name: str
    query_name: str
    index: int
    kind: str
    base: str


@dataclass
class TableMeta:
    key: str
    file_id: int
    file_path: str
    table_name: str
    row_count: int | None
    column_count: int
    columns: list[ColumnMeta]
    engine: str

    @property
    def estimated_cells(self) -> int | None:
        return self.row_count * self.column_count if self.row_count is not None else None


@dataclass
class FileMeta:
    id: int
    path: str
    absolute_path: str
    size: int
    mtime_ns: int
    sha256: str
    extension: str
    kind: str
    role_score: float
    tables: list[TableMeta] = field(default_factory=list)
    inventory_status: str = "ok"
    warning: str = ""


@dataclass
class ValueSource:
    fingerprint: str
    file_path: str
    table_name: str
    column: str
    kind: str
    base: str
    locator: str
    preview: str
    specificity: int
    origin: str

    def identity(self) -> tuple[str, str, str, str, str]:
        return self.file_path, self.table_name, self.column, self.locator, self.fingerprint


class DeadlineReached(RuntimeError):
    pass


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize_text(value: Any) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def normalize_value(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value)).casefold()


def value_fingerprint(value: Any) -> str:
    return hashlib.sha256(normalize_value(value).encode("utf-8")).hexdigest()


def safe_preview(kind: str, value: Any, limit: int = 72) -> str:
    text = normalize_text(value).replace("\r", " ").replace("\n", " ")
    if kind in {"name", "file_name", "column"}:
        return text[:limit] + ("..." if len(text) > limit else "")
    if kind == "email" and "@" in text:
        local, domain = text.split("@", 1)
        return f"{local[:1]}***@{domain}"
    if len(text) <= 4:
        return "*" * len(text)
    if len(text) <= 10:
        return f"{text[:1]}***{text[-1:]}"
    return f"{text[:4]}...{text[-4:]}"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def atomic_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temporary, path)


def goal_tokens(text: str) -> set[str]:
    normalized = re.sub(r"\s+", "", normalize_text(text).casefold())
    words = set(re.findall(r"[a-z0-9_]{2,}|[\u4e00-\u9fff]{2,}", normalized))
    chinese = "".join(re.findall(r"[\u4e00-\u9fff]", normalized))
    words.update(chinese[index:index + 2] for index in range(max(0, len(chinese) - 1)))
    return {word for word in words if word}


def normalized_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", normalize_text(value).casefold())


def header_semantic(value: Any) -> tuple[str, str]:
    text = normalized_header(value)
    if not text:
        return "other", ""
    if any(marker in text for marker in ("email", "邮箱", "邮件地址")):
        return "email", re.sub(r"email|邮箱|邮件地址", "", text)
    if any(marker in text for marker in ("phone", "mobile", "tel", "手机号", "联系电话", "电话")):
        return "phone", re.sub(r"phone|mobile|tel|手机号|联系电话|电话", "", text)
    if any(marker in text for marker in ("url", "uri", "网址", "链接")):
        return "url", re.sub(r"url|uri|网址|链接", "", text)
    reference_markers = ("parent", "source", "target", "from", "to", "父", "上级", "来源", "目标", "引用", "关联")
    id_markers = ("identifier", "uuid", "编号", "号码", "主键", "外键", "id")
    code_markers = ("code", "编码", "代码")
    name_markers = ("name", "title", "名称", "标题")
    if any(marker in text for marker in id_markers):
        base = text
        for marker in sorted(id_markers, key=len, reverse=True):
            base = base.replace(marker, "")
        return "reference" if any(marker in text for marker in reference_markers) else "id", base
    if any(marker in text for marker in code_markers):
        base = text
        for marker in code_markers:
            base = base.replace(marker, "")
        return "code", base
    if any(marker in text for marker in reference_markers):
        return "reference", text
    if any(marker in text for marker in name_markers):
        base = text
        for marker in name_markers:
            base = base.replace(marker, "")
        return "name", base
    return "other", text


def column_compatibility(left: ValueSource | ColumnMeta, right: ColumnMeta) -> float:
    left_name = normalized_header(left.column if isinstance(left, ValueSource) else left.name)
    right_name = normalized_header(right.name)
    if left_name and left_name == right_name:
        return 1.0
    if left.base and right.base and left.base == right.base:
        return 0.96
    if left.kind == right.kind and left.kind in {"email", "phone", "url"}:
        return 0.95
    if left.kind == right.kind and left.kind in {"id", "code", "reference"}:
        return 0.82
    if {left.kind, right.kind} <= {"id", "code", "reference"}:
        return 0.72
    if left.kind == right.kind == "name" and left.base and right.base:
        return 0.76
    return 0.0


def valid_join_value(value: Any, kind: str) -> str | None:
    text = normalize_value(value)
    if not text or len(text) > 256 or text in {"null", "none", "n/a", "na", "true", "false", "未知", "无"}:
        return None
    if re.fullmatch(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}(?:[t\s].*)?", text):
        return None
    if kind in {"id", "code", "reference", "phone"} and text.isdigit() and len(text) < 6:
        return None
    if kind == "name" and len(text) < 3:
        return None
    if kind == "other":
        return None
    return text


def specificity(value: str, kind: str) -> int:
    if kind in {"email", "phone", "url", "uuid"}:
        return 3
    if kind in {"id", "code", "reference", "labeled_value"}:
        if len(value) >= 8 or (any(char.isalpha() for char in value) and any(char.isdigit() for char in value)):
            return 3
        return 2
    return 2 if len(value) >= 8 else 1


class BottomK:
    def __init__(self, limit: int, items: Iterable[ValueSource] | None = None) -> None:
        self.limit = limit
        self.items: dict[str, ValueSource] = {}
        for item in items or ():
            self.add(item)

    def add(self, source: ValueSource) -> None:
        current = self.items.get(source.fingerprint)
        if current is not None:
            return
        self.items[source.fingerprint] = source
        if len(self.items) > self.limit:
            del self.items[max(self.items)]

    def values(self) -> list[ValueSource]:
        return [self.items[key] for key in sorted(self.items)]


class CardinalitySketch:
    """Small HyperLogLog sketch used only to suppress low-cardinality expansion."""

    def __init__(self, nonempty: int = 0, registers: Sequence[int] | None = None) -> None:
        self.nonempty = nonempty
        self.registers = list(registers) if registers is not None else [0] * 256

    def add(self, fingerprint: str) -> None:
        value = int(fingerprint[:16], 16)
        bucket = value & 0xFF
        remainder = value >> 8
        rank = 57 if remainder == 0 else 57 - remainder.bit_length()
        self.registers[bucket] = max(self.registers[bucket], rank)
        self.nonempty += 1

    def estimate(self) -> float:
        size = len(self.registers)
        raw = 0.7213 / (1 + 1.079 / size) * size * size / sum(2.0 ** -value for value in self.registers)
        zeroes = self.registers.count(0)
        if zeroes and raw <= 2.5 * size:
            import math
            return size * math.log(size / zeroes)
        return raw

    def ratio(self) -> float:
        return min(1.0, self.estimate() / self.nonempty) if self.nonempty else 0.0

    def to_json(self) -> dict[str, Any]:
        return {"nonempty": self.nonempty, "registers": self.registers}


class SeedIndex:
    def __init__(self, max_values: int, max_sources_per_value: int = 4) -> None:
        self.max_values = max_values
        self.max_sources_per_value = max_sources_per_value
        self.by_fingerprint: dict[str, list[ValueSource]] = {}

    def add(self, source: ValueSource) -> bool:
        sources = self.by_fingerprint.get(source.fingerprint)
        if sources is None:
            if len(self.by_fingerprint) >= self.max_values:
                return False
            sources = self.by_fingerprint[source.fingerprint] = []
        if source.identity() in {item.identity() for item in sources}:
            return False
        if len(sources) >= self.max_sources_per_value:
            return False
        sources.append(source)
        return True

    def matching(self, source: ValueSource) -> list[tuple[ValueSource, float]]:
        matches: list[tuple[ValueSource, float]] = []
        target_column = ColumnMeta(source.column, source.column, 0, source.kind, source.base)
        for existing in self.by_fingerprint.get(source.fingerprint, []):
            if existing.file_path == source.file_path:
                continue
            score = column_compatibility(existing, target_column)
            if score >= 0.70 or min(existing.specificity, source.specificity) >= 3:
                matches.append((existing, max(score, 0.78)))
        return matches

    def signatures(self) -> set[tuple[str, str]]:
        return {(source.kind, source.base) for sources in self.by_fingerprint.values() for source in sources}

    def to_json(self) -> list[dict[str, Any]]:
        return [asdict(source) for sources in self.by_fingerprint.values() for source in sources]

    @classmethod
    def from_json(cls, payload: list[dict[str, Any]], max_values: int) -> "SeedIndex":
        instance = cls(max_values)
        for item in payload:
            instance.add(ValueSource(**item))
        return instance


class RelationBook:
    def __init__(self, relations: Iterable[dict[str, Any]] | None = None) -> None:
        self.items: dict[str, dict[str, Any]] = {}
        for relation in relations or ():
            self.items[relation["id"]] = relation

    def add(
        self,
        source: ValueSource,
        target: ValueSource,
        relation_type: str,
        confidence: float,
        verdict: str,
        evidence_kind: str,
        note: str,
    ) -> None:
        if source.file_path == target.file_path:
            return
        if source.file_path > target.file_path:
            source, target = target, source
        key = "\0".join((source.file_path, target.file_path, relation_type, source.column, target.column))
        relation_id = "R-" + hashlib.sha1(key.encode("utf-8")).hexdigest()[:12]
        evidence = {
            "kind": evidence_kind,
            "source_locator": source.locator,
            "target_locator": target.locator,
            "source_column": source.column,
            "target_column": target.column,
            "value_preview": source.preview if source.specificity >= target.specificity else target.preview,
            "fingerprint": source.fingerprint,
            "strength": "strong" if min(source.specificity, target.specificity) >= 2 else "corroborating",
            "note": note,
        }
        relation = self.items.get(relation_id)
        evidence_identity = (evidence["fingerprint"], evidence["source_locator"], evidence["target_locator"])
        if relation is None:
            relation = self.items[relation_id] = {
                "id": relation_id,
                "source": source.file_path,
                "target": target.file_path,
                "type": relation_type,
                "direction": "undirected",
                "verdict": verdict,
                "confidence": round(confidence, 3),
                "source_column": source.column,
                "target_column": target.column,
                "explanation": note,
                "evidence_count": 0,
                "evidence": [],
                "omitted_evidence_count": 0,
            }
        if evidence_identity in {
            (item["fingerprint"], item["source_locator"], item["target_locator"])
            for item in relation["evidence"]
        }:
            return
        relation["evidence_count"] += 1
        if len(relation["evidence"]) < MAX_EVIDENCE:
            relation["evidence"].append(evidence)
        else:
            relation["omitted_evidence_count"] += 1
        relation["confidence"] = round(max(relation["confidence"], confidence), 3)
        if verdict == "confirmed":
            relation["verdict"] = "confirmed"

    def add_duplicate(self, left: FileMeta, right: FileMeta) -> None:
        source = ValueSource(left.sha256, left.path, "", "whole-file", "id", "file", "whole-file", left.sha256[:16] + "...", 3, "hash")
        target = ValueSource(right.sha256, right.path, "", "whole-file", "id", "file", "whole-file", right.sha256[:16] + "...", 3, "hash")
        self.add(source, target, "exact_duplicate", 1.0, "confirmed", "sha256", "Complete file SHA-256 digests are identical.")

    def values(self) -> list[dict[str, Any]]:
        return sorted(self.items.values(), key=lambda item: (item["source"], item["target"], item["type"], item["source_column"]))


def detect_encoding(path: Path) -> str:
    with path.open("rb") as stream:
        sample = stream.read(65_536)
    if sample.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if sample.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16"
    for encoding in ("utf-8", "gb18030"):
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            pass
    return "utf-8"


def role_score(path: Path, goal: set[str]) -> float:
    name = normalized_header(path.stem)
    score = sum(1.0 for term in ROLE_TERMS if normalized_header(term) in name)
    file_tokens = goal_tokens(path.stem)
    if goal and file_tokens:
        score += 2.0 * len(goal & file_tokens) / max(1, len(file_tokens))
    return score


def make_columns(headers: Sequence[Any]) -> list[ColumnMeta]:
    columns: list[ColumnMeta] = []
    counts: dict[str, int] = defaultdict(int)
    for index, raw in enumerate(headers):
        name = normalize_text(raw) or f"column_{index + 1}"
        counts[name] += 1
        query_name = name if counts[name] == 1 else f"{name}_{counts[name] - 1}"
        kind, base = header_semantic(name)
        columns.append(ColumnMeta(name, query_name, index, kind, base))
    return columns


def _xlsx_layouts(path: Path) -> list[tuple[str, str, int | None, int | None, str]]:
    with zipfile.ZipFile(path) as archive:
        workbook = ElementTree.parse(archive.open("xl/workbook.xml")).getroot()
        relationships = ElementTree.parse(archive.open("xl/_rels/workbook.xml.rels")).getroot()
        targets = {element.attrib.get("Id", ""): element.attrib.get("Target", "") for element in relationships}
        layouts: list[tuple[str, str, int | None, int | None, str]] = []
        relation_attribute = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        for sheet in workbook.iter():
            if not sheet.tag.endswith("}sheet"):
                continue
            name = sheet.attrib.get("name", "Sheet")
            target = targets.get(sheet.attrib.get(relation_attribute, ""), "")
            entry = target.lstrip("/")
            if not entry.startswith("xl/"):
                entry = "xl/" + entry
            dimension = ""
            try:
                with archive.open(entry) as stream:
                    for _event, element in ElementTree.iterparse(stream, events=("start",)):
                        if element.tag.endswith("}dimension"):
                            dimension = element.attrib.get("ref", "")
                            break
            except KeyError:
                pass
            row_count, column_count, end_column = parse_dimension(dimension)
            layouts.append((name, entry, row_count, column_count, end_column))
        return layouts


def parse_dimension(dimension: str) -> tuple[int | None, int | None, str]:
    end = dimension.split(":")[-1]
    match = re.fullmatch(r"([A-Z]+)(\d+)", end, re.IGNORECASE)
    if not match:
        return None, None, "Z"
    letters, rows = match.groups()
    column_number = 0
    for char in letters.upper():
        column_number = column_number * 26 + ord(char) - 64
    return int(rows), column_number, letters.upper()


def _duckdb() -> Any | None:
    try:
        import duckdb
    except ImportError:
        return None
    return duckdb


def _fastexcel() -> Any | None:
    try:
        import fastexcel
    except ImportError:
        return None
    return fastexcel


def _fastexcel_xlsx_headers(path: Path, sheet: str) -> list[str]:
    fastexcel = _fastexcel()
    if fastexcel is None:
        raise RuntimeError("fastexcel unavailable")
    reader = fastexcel.read_excel(str(path))
    batch = reader.load_sheet_eager(sheet, header_row=0, n_rows=1, dtypes="string", dtype_coercion="coerce")
    return list(batch.schema.names)


def _duckdb_xlsx_headers(path: Path, sheet: str, end_column: str) -> list[str]:
    duckdb = _duckdb()
    if duckdb is None:
        raise RuntimeError("duckdb unavailable")
    connection = duckdb.connect()
    try:
        query = "SELECT * FROM read_xlsx(?, sheet=?, range=?, header=true, all_varchar=true, ignore_errors=true) LIMIT 0"
        cursor = connection.execute(query, [str(path), sheet, f"A1:{end_column}2"])
        return [item[0] for item in cursor.description]
    finally:
        connection.close()


def inspect_xlsx(file_id: int, relative: str, path: Path) -> list[TableMeta]:
    layouts = _xlsx_layouts(path)
    tables: list[TableMeta] = []
    fallback_headers: dict[str, list[Any]] = {}
    for sheet, _entry, rows, columns, end_column in layouts:
        engine = "fastexcel"
        try:
            headers = _fastexcel_xlsx_headers(path, sheet)
        except Exception:
            engine = "duckdb"
            try:
                headers = _duckdb_xlsx_headers(path, sheet, end_column)
            except Exception:
                engine = "openpyxl"
                if not fallback_headers:
                    try:
                        import openpyxl
                    except ImportError as exc:
                        raise RuntimeError("XLSX requires fastexcel, duckdb, or openpyxl") from exc
                    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    try:
                        for worksheet in workbook.worksheets:
                            first = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                            fallback_headers[worksheet.title] = list(first)
                    finally:
                        workbook.close()
                headers = [normalize_text(value) or f"column_{index + 1}" for index, value in enumerate(fallback_headers.get(sheet, []))]
        table_columns = make_columns(headers)
        tables.append(TableMeta(f"{file_id}:{sheet}", file_id, relative, sheet, rows, columns or len(table_columns), table_columns, engine))
    return tables


def inspect_csv(file_id: int, relative: str, path: Path) -> list[TableMeta]:
    encoding = detect_encoding(path)
    with path.open("r", encoding=encoding, errors="replace", newline="") as stream:
        if path.suffix.casefold() == ".tsv":
            reader = csv.reader(stream, dialect="excel-tab")
        else:
            sample = stream.read(8192)
            stream.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = "excel"
            reader = csv.reader(stream, dialect=dialect)
        headers = next(reader, [])
    columns = make_columns(headers)
    return [TableMeta(f"{file_id}:{path.name}", file_id, relative, path.name, None, len(columns), columns, "csv")]


def inspect_jsonl(file_id: int, relative: str, path: Path) -> list[TableMeta]:
    headers: list[str] = []
    seen: set[str] = set()
    with path.open("r", encoding=detect_encoding(path), errors="replace") as stream:
        for line in itertools.islice(stream, 100):
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(item, dict):
                for key in item:
                    if key not in seen:
                        seen.add(key)
                        headers.append(str(key))
    columns = make_columns(headers)
    return [TableMeta(f"{file_id}:{path.name}", file_id, relative, path.name, None, len(columns), columns, "jsonl")]


def inspect_sqlite(file_id: int, relative: str, path: Path) -> list[TableMeta]:
    connection = sqlite3.connect(path.resolve().as_uri() + "?mode=ro", uri=True)
    tables: list[TableMeta] = []
    try:
        names = connection.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'").fetchall()
        for (name,) in names:
            escaped = '"' + name.replace('"', '""') + '"'
            headers = [row[1] for row in connection.execute(f"PRAGMA table_info({escaped})")]
            columns = make_columns(headers)
            tables.append(TableMeta(f"{file_id}:{name}", file_id, relative, name, None, len(columns), columns, "sqlite"))
    finally:
        connection.close()
    return tables


def inspect_parquet(file_id: int, relative: str, path: Path) -> list[TableMeta]:
    duckdb = _duckdb()
    if duckdb is None:
        raise RuntimeError("Parquet requires duckdb")
    connection = duckdb.connect()
    try:
        cursor = connection.execute("SELECT * FROM read_parquet(?) LIMIT 0", [str(path)])
        headers = [item[0] for item in cursor.description]
        row_count = connection.execute("SELECT count(*) FROM read_parquet(?)", [str(path)]).fetchone()[0]
    finally:
        connection.close()
    columns = make_columns(headers)
    return [TableMeta(f"{file_id}:{path.name}", file_id, relative, path.name, row_count, len(columns), columns, "duckdb-parquet")]


def inventory(input_root: Path, goal: set[str]) -> tuple[list[FileMeta], list[str]]:
    warnings: list[str] = []
    paths = sorted((path for path in input_root.rglob("*") if path.is_file() and not path.is_symlink()), key=lambda path: path.relative_to(input_root).as_posix().casefold())
    files: list[FileMeta] = []
    for file_id, path in enumerate(paths, 1):
        relative = path.relative_to(input_root).as_posix()
        extension = path.suffix.casefold()
        stat = path.stat()
        kind = "tabular" if extension in TABULAR_EXTENSIONS else "document" if extension in DOCUMENT_EXTENSIONS or extension in OCR_EXTENSIONS else "binary"
        item = FileMeta(file_id, relative, str(path), stat.st_size, stat.st_mtime_ns, sha256_file(path), extension, kind, role_score(path, goal))
        try:
            if extension in {".csv", ".tsv"}:
                item.tables = inspect_csv(file_id, relative, path)
            elif extension == ".xlsx":
                item.tables = inspect_xlsx(file_id, relative, path)
            elif extension in {".jsonl", ".ndjson"}:
                item.tables = inspect_jsonl(file_id, relative, path)
            elif extension in {".sqlite", ".sqlite3", ".db"}:
                item.tables = inspect_sqlite(file_id, relative, path)
            elif extension == ".parquet":
                item.tables = inspect_parquet(file_id, relative, path)
        except Exception as exc:
            item.inventory_status = "metadata_only"
            item.warning = f"{type(exc).__name__}: {exc}"
            warnings.append(f"Inventory failed for {relative}: {item.warning}")
        files.append(item)
    return files, warnings


def _quoted_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def iter_table_rows(table: TableMeta, path: Path, selected: Sequence[ColumnMeta], start_offset: int = 0) -> Iterator[tuple[int, dict[str, Any]]]:
    if not selected:
        return
    if table.engine == "fastexcel":
        fastexcel = _fastexcel()
        if fastexcel is not None:
            reader = fastexcel.read_excel(str(path))
            batch = reader.load_sheet_eager(
                table.table_name,
                header_row=0,
                use_columns=[column.index for column in selected],
                dtypes="string",
                dtype_coercion="coerce",
            )
            for batch_offset in range(start_offset, batch.num_rows, 4096):
                records = batch.slice(batch_offset, min(4096, batch.num_rows - batch_offset)).to_pylist()
                for record_offset, record in enumerate(records):
                    row_number = batch_offset + record_offset + 2
                    yield row_number, {
                        column.name: record.get(batch.schema.names[index])
                        for index, column in enumerate(selected)
                    }
            return
    if table.engine == "duckdb":
        duckdb = _duckdb()
        if duckdb is not None:
            connection = duckdb.connect()
            try:
                projection = ", ".join(_quoted_identifier(column.query_name) for column in selected)
                query = (
                    f"SELECT row_number() OVER () + 1 AS __row_number, {projection} "
                    "FROM read_xlsx(?, sheet=?, header=true, all_varchar=true, ignore_errors=true) "
                    f"OFFSET {int(start_offset)}"
                )
                cursor = connection.execute(query, [str(path), table.table_name])
                while rows := cursor.fetchmany(4096):
                    for row in rows:
                        yield int(row[0]), {column.name: row[index + 1] for index, column in enumerate(selected)}
                return
            except Exception:
                pass
            finally:
                connection.close()
    if table.engine in {"duckdb-parquet"}:
        duckdb = _duckdb()
        if duckdb is None:
            raise RuntimeError("duckdb unavailable")
        connection = duckdb.connect()
        try:
            projection = ", ".join(_quoted_identifier(column.query_name) for column in selected)
            cursor = connection.execute(
                f"SELECT row_number() OVER () AS __row_number, {projection} FROM read_parquet(?) OFFSET {int(start_offset)}",
                [str(path)],
            )
            while rows := cursor.fetchmany(4096):
                for row in rows:
                    yield int(row[0]), {column.name: row[index + 1] for index, column in enumerate(selected)}
        finally:
            connection.close()
        return
    if path.suffix.casefold() == ".xlsx":
        import openpyxl
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[table.table_name]
            for data_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 1):
                if data_index <= start_offset:
                    continue
                yield data_index + 1, {column.name: row[column.index] if column.index < len(row) else None for column in selected}
        finally:
            workbook.close()
        return
    if path.suffix.casefold() in {".csv", ".tsv"}:
        encoding = detect_encoding(path)
        with path.open("r", encoding=encoding, errors="replace", newline="") as stream:
            if path.suffix.casefold() == ".tsv":
                reader = csv.reader(stream, dialect="excel-tab")
            else:
                sample = stream.read(8192)
                stream.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = "excel"
                reader = csv.reader(stream, dialect=dialect)
            next(reader, None)
            for data_index, row in enumerate(reader, 1):
                if data_index <= start_offset:
                    continue
                yield data_index + 1, {column.name: row[column.index] if column.index < len(row) else None for column in selected}
        return
    if path.suffix.casefold() in {".jsonl", ".ndjson"}:
        with path.open("r", encoding=detect_encoding(path), errors="replace") as stream:
            for data_index, line in enumerate(stream, 1):
                if data_index <= start_offset:
                    continue
                try:
                    item = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if isinstance(item, dict):
                    yield data_index, {column.name: item.get(column.name) for column in selected}
        return
    if path.suffix.casefold() in {".sqlite", ".sqlite3", ".db"}:
        connection = sqlite3.connect(path.resolve().as_uri() + "?mode=ro", uri=True)
        escaped_table = _quoted_identifier(table.table_name)
        projection = ", ".join(_quoted_identifier(column.name) for column in selected)
        try:
            cursor = connection.execute(f"SELECT rowid, {projection} FROM {escaped_table} LIMIT -1 OFFSET ?", (start_offset,))
            for row in cursor:
                yield int(row[0]), {column.name: row[index + 1] for index, column in enumerate(selected)}
        finally:
            connection.close()


UUID_RE = re.compile(r"(?i)(?<![0-9a-f])[0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}(?![0-9a-f])")
EMAIL_RE = re.compile(r"(?i)(?<![\w.+-])[\w.+-]{1,64}@[a-z0-9-]+(?:\.[a-z0-9-]+)+(?![\w.-])")
PHONE_RE = re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)")
URL_RE = re.compile(r"(?i)https?://[^\s<>\"']{6,300}")
LABELED_RE = re.compile(r"(?P<label>[A-Za-z\u4e00-\u9fff][A-Za-z0-9_\u4e00-\u9fff -]{1,30})\s*[:：#]\s*(?P<value>[A-Za-z0-9][A-Za-z0-9_./-]{3,80})")


def extract_text_values(text: str, locator: str, file_path: str) -> Iterator[ValueSource]:
    patterns = (("uuid", UUID_RE), ("email", EMAIL_RE), ("phone", PHONE_RE), ("url", URL_RE))
    for kind, pattern in patterns:
        for match in pattern.finditer(text):
            value = normalize_value(match.group(0).rstrip(".,;:!?，。；：！？)"))
            yield ValueSource(value_fingerprint(value), file_path, "", kind, kind, kind, locator, safe_preview(kind, value), 3, "document")
    for match in LABELED_RE.finditer(text):
        label, value = match.group("label"), normalize_value(match.group("value"))
        header_kind, base = header_semantic(label)
        kind = header_kind if header_kind != "other" else "labeled_value"
        yield ValueSource(value_fingerprint(value), file_path, "", label, kind, base, locator, safe_preview(kind, value), specificity(value, kind), "document")


def iter_document_segments(path: Path, ocr_mode: str) -> Iterator[tuple[str, str]]:
    extension = path.suffix.casefold()
    if extension in TEXT_EXTENSIONS:
        with path.open("r", encoding=detect_encoding(path), errors="replace") as stream:
            for line_number, line in enumerate(stream, 1):
                yield f"line:{line_number}", line
        return
    if extension == ".docx":
        with zipfile.ZipFile(path) as archive, archive.open("word/document.xml") as stream:
            paragraph = 0
            for _event, element in ElementTree.iterparse(stream, events=("end",)):
                if element.tag.endswith("}p"):
                    paragraph += 1
                    yield f"paragraph:{paragraph}", "".join(node.text or "" for node in element.iter() if node.tag.endswith("}t"))
                    element.clear()
        return
    if extension == ".pptx":
        with zipfile.ZipFile(path) as archive:
            names = sorted(name for name in archive.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", name))
            for index, name in enumerate(names, 1):
                root = ElementTree.parse(archive.open(name)).getroot()
                yield f"slide:{index}", " ".join(node.text or "" for node in root.iter() if node.tag.endswith("}t"))
        return
    if extension == ".pdf" and ocr_mode != "always":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            extracted = 0
            for page_number, page in enumerate(reader.pages, 1):
                text = page.extract_text() or ""
                extracted += len(text)
                yield f"page:{page_number}", text
            if extracted >= 500 or ocr_mode == "never":
                return
        except Exception:
            if ocr_mode == "never":
                return
    if extension in OCR_EXTENSIONS and ocr_mode != "never":
        yield from iter_ocr_segments(path)


def find_ocr_script() -> Path | None:
    script = Path(__file__).resolve().parents[2] / "ocr-parser" / "scripts" / "parse.py"
    return script if script.is_file() else None


def iter_ocr_segments(path: Path) -> Iterator[tuple[str, str]]:
    script = find_ocr_script()
    if script is None:
        raise RuntimeError("ocr-parser Skill is unavailable")
    with tempfile.TemporaryDirectory(prefix="discover-data-relations-ocr-") as temporary:
        output_path = Path(temporary) / "ocr.txt"
        with output_path.open("wb") as output:
            completed = subprocess.run(
                [sys.executable, str(script), "--path", str(path), "--format", "text"],
                stdin=subprocess.DEVNULL,
                stdout=output,
                stderr=subprocess.DEVNULL,
                timeout=900,
                check=False,
            )
        if completed.returncode != 0:
            raise RuntimeError(f"ocr-parser exited with {completed.returncode}")
        with output_path.open("r", encoding=detect_encoding(output_path), errors="replace") as stream:
            for line_number, line in enumerate(stream, 1):
                yield f"ocr:line:{line_number}", line


def source_from_cell(file_path: str, table: TableMeta, column: ColumnMeta, row_number: int, value: Any, origin: str) -> ValueSource | None:
    normalized = valid_join_value(value, column.kind)
    if normalized is None:
        return None
    return ValueSource(
        value_fingerprint(normalized), file_path, table.table_name, column.name, column.kind, column.base,
        f"table:{table.table_name};row:{row_number};column:{column.name}", safe_preview(column.kind, normalized),
        specificity(normalized, column.kind), origin,
    )


class ProgressiveAnalyzer:
    def __init__(self, input_root: Path, output_root: Path, args: argparse.Namespace) -> None:
        self.input_root = input_root.resolve()
        self.output_root = output_root.resolve()
        self.args = args
        self.deadline = time.monotonic() + args.deadline_seconds if args.deadline_seconds > 0 else float("inf")
        self.files: list[FileMeta] = []
        self.file_by_id: dict[int, FileMeta] = {}
        self.tables: dict[str, TableMeta] = {}
        self.seeds = SeedIndex(args.max_seed_values)
        self.relations = RelationBook()
        self.profiles: dict[str, BottomK] = {}
        self.frontier_profiles: dict[str, BottomK] = {}
        self.column_stats: dict[str, CardinalitySketch] = {}
        self.match_counts: dict[str, dict[str, int]] = {}
        self.scanned_tables: set[str] = set()
        self.partial_table: dict[str, Any] | None = None
        self.scan_stats: dict[str, dict[str, Any]] = {}
        self.warnings: list[str] = []
        self.signature = ""
        self.resumed = False

    def check_deadline(self) -> None:
        if time.monotonic() >= self.deadline:
            raise DeadlineReached

    def load_or_initialize(self, goal: set[str]) -> None:
        self.files, inventory_warnings = inventory(self.input_root, goal)
        self.warnings.extend(inventory_warnings)
        self.file_by_id = {item.id: item for item in self.files}
        self.tables = {table.key: table for item in self.files for table in item.tables}
        signature_input = "\n".join(f"{item.path}\0{item.size}\0{item.mtime_ns}\0{item.sha256}" for item in self.files)
        self.signature = hashlib.sha256(signature_input.encode("utf-8")).hexdigest()
        catalog = {
            "schema_version": SCHEMA_VERSION,
            "input_root": str(self.input_root),
            "signature": self.signature,
            "files": [self.file_json(item) for item in self.files],
        }
        atomic_json(self.output_root / "catalog.json", catalog)
        state_path = self.output_root / "progress.json"
        if self.args.resume and state_path.is_file():
            try:
                state = json.loads(state_path.read_text(encoding="utf-8"))
                if state.get("schema_version") == SCHEMA_VERSION and state.get("signature") == self.signature:
                    self.seeds = SeedIndex.from_json(state.get("seeds", []), self.args.max_seed_values)
                    self.relations = RelationBook(state.get("relations", []))
                    self.profiles = {
                        key: BottomK(self.args.profile_size, [ValueSource(**item) for item in values])
                        for key, values in state.get("profiles", {}).items()
                    }
                    self.frontier_profiles = {
                        key: BottomK(self.args.frontier_values_per_column, [ValueSource(**item) for item in values])
                        for key, values in state.get("frontier_profiles", {}).items()
                    }
                    self.column_stats = {
                        key: CardinalitySketch(int(value.get("nonempty", 0)), value.get("registers", []))
                        for key, value in state.get("column_stats", {}).items()
                    }
                    self.match_counts = {
                        key: {fingerprint: int(count) for fingerprint, count in values.items()}
                        for key, values in state.get("match_counts", {}).items()
                    }
                    self.scanned_tables = set(state.get("scanned_tables", []))
                    self.partial_table = state.get("partial_table")
                    self.scan_stats = state.get("scan_stats", {})
                    self.warnings = list(dict.fromkeys(self.warnings + state.get("warnings", [])))
                    self.resumed = True
            except (OSError, ValueError, TypeError, json.JSONDecodeError):
                pass
        if not self.resumed:
            self.cleanup_legacy_index()
            groups: dict[str, list[FileMeta]] = defaultdict(list)
            for item in self.files:
                groups[item.sha256].append(item)
            for matches in groups.values():
                for left, right in itertools.combinations(matches, 2):
                    self.relations.add_duplicate(left, right)

    def cleanup_legacy_index(self) -> None:
        for name in ("evidence.sqlite3", "evidence.sqlite3-wal", "evidence.sqlite3-shm"):
            path = self.output_root / name
            if path.exists():
                path.unlink()

    @staticmethod
    def file_json(item: FileMeta) -> dict[str, Any]:
        payload = asdict(item)
        payload.pop("absolute_path", None)
        return payload

    def save_progress(self) -> None:
        payload = {
            "schema_version": SCHEMA_VERSION,
            "signature": self.signature,
            "updated_at": utc_now(),
            "seeds": self.seeds.to_json(),
            "relations": self.relations.values(),
            "profiles": {key: [asdict(item) for item in profile.values()] for key, profile in self.profiles.items()},
            "frontier_profiles": {key: [asdict(item) for item in profile.values()] for key, profile in self.frontier_profiles.items()},
            "column_stats": {key: value.to_json() for key, value in self.column_stats.items()},
            "match_counts": self.match_counts,
            "scanned_tables": sorted(self.scanned_tables),
            "partial_table": self.partial_table,
            "scan_stats": self.scan_stats,
            "warnings": list(dict.fromkeys(self.warnings)),
        }
        atomic_json(self.output_root / "progress.json", payload)

    def connect_and_add(self, source: ValueSource, relation_type: str = "seeded_value_link") -> bool:
        matched = False
        for existing, compatibility in self.seeds.matching(source):
            minimum_specificity = min(existing.specificity, source.specificity)
            confirmed = (
                compatibility >= 0.95 and minimum_specificity >= 2
            ) or (
                compatibility >= 0.82 and minimum_specificity >= 3
            )
            confidence = min(0.99, 0.82 + 0.12 * compatibility + 0.015 * min(existing.specificity, source.specificity))
            self.relations.add(
                existing, source, relation_type, confidence,
                "confirmed" if confirmed else "supported_hypothesis", "exact_value_match",
                f"Exact normalized values match across compatible columns (column compatibility {compatibility:.2f}).",
            )
            matched = True
        self.seeds.add(source)
        return matched

    def seed_candidates(self) -> tuple[list[TableMeta], list[FileMeta]]:
        tables: list[TableMeta] = []
        documents: list[FileMeta] = []
        for item in self.files:
            for table in item.tables:
                cells = table.estimated_cells
                if cells is not None and cells <= self.args.seed_cell_budget:
                    tables.append(table)
                elif cells is None and item.size <= self.args.seed_file_bytes:
                    tables.append(table)
            if item.kind == "document" and item.size <= self.args.seed_file_bytes:
                documents.append(item)
        tables.sort(key=lambda table: (-self.file_by_id[table.file_id].role_score, table.estimated_cells or self.file_by_id[table.file_id].size))
        documents.sort(key=lambda item: (-item.role_score, item.size))
        if not tables and self.tables:
            tables = [min(self.tables.values(), key=lambda table: table.estimated_cells or self.file_by_id[table.file_id].size)]
        return tables, documents

    def process_seed_table(self, table: TableMeta, bootstrap: bool = False) -> None:
        if table.key in self.scanned_tables and not bootstrap:
            return
        columns = table.columns
        join_columns = [column for column in columns if column.kind in JOIN_KINDS]
        reservoirs = {column.name: BottomK(self.args.seed_values_per_column) for column in join_columns}
        text_reservoir = BottomK(self.args.max_text_seeds)
        path = Path(self.file_by_id[table.file_id].absolute_path)
        rows = 0
        for row_number, values in iter_table_rows(table, path, columns):
            rows += 1
            if bootstrap and rows > self.args.bootstrap_rows:
                break
            for column in join_columns:
                source = source_from_cell(table.file_path, table, column, row_number, values.get(column.name), "seed_table")
                if source:
                    reservoirs[column.name].add(source)
                    self.profile_for(table, column).add(source)
            if not bootstrap:
                combined = "\t".join(normalize_text(value) for value in values.values() if value not in (None, ""))
                for source in extract_text_values(combined, f"table:{table.table_name};row:{row_number}", table.file_path):
                    text_reservoir.add(source)
        for reservoir in reservoirs.values():
            for source in reservoir.values():
                self.connect_and_add(source)
        for source in text_reservoir.values():
            self.connect_and_add(source)
        self.scan_stats[table.key] = {
            "mode": "bootstrap_seed" if bootstrap else "complete_seed",
            "rows_scanned": rows,
            "candidate_columns": [column.name for column in join_columns],
            "matches": 0,
            "engine": table.engine,
        }
        if not bootstrap:
            self.scanned_tables.add(table.key)
        self.save_progress()

    def process_seed_document(self, item: FileMeta, all_names: Sequence[str]) -> None:
        key = f"document:{item.id}"
        if key in self.scanned_tables:
            return
        reservoir = BottomK(self.args.max_text_seeds)
        characters = 0
        try:
            for locator, text in iter_document_segments(Path(item.absolute_path), self.args.ocr_mode):
                characters += len(text)
                for source in extract_text_values(text, locator, item.path):
                    reservoir.add(source)
                folded = normalize_text(text).casefold()
                for target_name in all_names:
                    if target_name != item.path and len(Path(target_name).name) >= 4 and Path(target_name).name.casefold() in folded:
                        source = ValueSource(value_fingerprint(target_name), item.path, "", "file-reference", "reference", "file", locator, Path(target_name).name, 3, "document")
                        target = ValueSource(value_fingerprint(target_name), target_name, "", "file-name", "reference", "file", "file-name", Path(target_name).name, 3, "file")
                        self.relations.add(source, target, "explicit_reference", 0.99, "confirmed", "explicit_file_reference", "The source text explicitly names the target file.")
                if characters >= self.args.document_character_budget:
                    self.warnings.append(f"Document seed budget reached for {item.path}")
                    break
        except Exception as exc:
            self.warnings.append(f"Document parsing failed for {item.path}: {type(exc).__name__}: {exc}")
        for source in reservoir.values():
            self.connect_and_add(source)
        self.scan_stats[key] = {"mode": "document_seed", "characters_scanned": characters, "matches": 0, "engine": "document"}
        self.scanned_tables.add(key)
        self.save_progress()

    def profile_for(self, table: TableMeta, column: ColumnMeta) -> BottomK:
        key = f"{table.key}\0{column.name}"
        if key not in self.profiles:
            self.profiles[key] = BottomK(self.args.profile_size)
        return self.profiles[key]

    def frontier_for(self, table: TableMeta, column: ColumnMeta) -> BottomK:
        key = f"{table.key}\0{column.name}"
        if key not in self.frontier_profiles:
            self.frontier_profiles[key] = BottomK(self.args.frontier_values_per_column)
        return self.frontier_profiles[key]

    def stats_for(self, table: TableMeta, column: ColumnMeta) -> CardinalitySketch:
        key = f"{table.key}\0{column.name}"
        if key not in self.column_stats:
            self.column_stats[key] = CardinalitySketch()
        return self.column_stats[key]

    def column_can_reach_unscanned_table(self, table: TableMeta, column: ColumnMeta) -> bool:
        synthetic = ValueSource("", table.file_path, table.table_name, column.name, column.kind, column.base, "", "", 0, "")
        return any(
            other.file_path != table.file_path
            and other.key not in self.scanned_tables
            and any(column_compatibility(synthetic, candidate) >= 0.72 for candidate in other.columns)
            for other in self.tables.values()
        )

    def expand_completed_table(self, table: TableMeta, selected: Sequence[ColumnMeta]) -> None:
        for column in selected:
            if not self.column_can_reach_unscanned_table(table, column):
                continue
            stats = self.stats_for(table, column)
            if stats.ratio() < self.args.min_expansion_distinct_ratio:
                continue
            profile = self.frontier_profiles.get(f"{table.key}\0{column.name}")
            if profile:
                for source in profile.values():
                    self.seeds.add(source)

    def table_score(self, table: TableMeta) -> tuple[float, float]:
        signatures = self.seeds.signatures()
        score = 0.0
        for column in table.columns:
            for kind, base in signatures:
                synthetic = ValueSource("", "", "", "", kind, base, "", "", 0, "")
                score = max(score, column_compatibility(synthetic, column))
        size = table.estimated_cells or self.file_by_id[table.file_id].size
        return score, -float(size)

    def scan_large_table(self, table: TableMeta, start_offset: int = 0) -> None:
        selected = [column for column in table.columns if column.kind in JOIN_KINDS]
        if not selected:
            self.scanned_tables.add(table.key)
            self.scan_stats[table.key] = {"mode": "schema_only", "rows_scanned": 0, "candidate_columns": [], "matches": 0, "engine": table.engine}
            self.partial_table = None
            self.save_progress()
            return
        path = Path(self.file_by_id[table.file_id].absolute_path)
        rows_scanned = start_offset
        matches = int(self.scan_stats.get(table.key, {}).get("matches", 0))
        table_match_counts = self.match_counts.setdefault(table.key, {})
        for row_number, values in iter_table_rows(table, path, selected, start_offset):
            rows_scanned += 1
            row_sources: list[ValueSource] = []
            row_matched = False
            for column in selected:
                source = source_from_cell(table.file_path, table, column, row_number, values.get(column.name), "large_probe")
                if source is None:
                    continue
                row_sources.append(source)
                self.profile_for(table, column).add(source)
                self.stats_for(table, column).add(source.fingerprint)
                if table_match_counts.get(source.fingerprint, 0) >= self.args.max_matches_per_seed:
                    continue
                for existing, compatibility in self.seeds.matching(source):
                    minimum_specificity = min(existing.specificity, source.specificity)
                    confirmed = (
                        compatibility >= 0.95 and minimum_specificity >= 2
                    ) or (
                        compatibility >= 0.82 and minimum_specificity >= 3
                    )
                    confidence = min(0.99, 0.82 + 0.12 * compatibility + 0.015 * min(existing.specificity, source.specificity))
                    self.relations.add(existing, source, "seeded_value_link", confidence, "confirmed" if confirmed else "supported_hypothesis", "exact_value_match", f"A bounded seed exactly matches this candidate column (column compatibility {compatibility:.2f}).")
                    row_matched = True
                    matches += 1
                    table_match_counts[source.fingerprint] = table_match_counts.get(source.fingerprint, 0) + 1
            if row_matched:
                columns_by_name = {column.name: column for column in selected}
                for source in row_sources:
                    self.frontier_for(table, columns_by_name[source.column]).add(source)
            if rows_scanned % self.args.checkpoint_rows == 0:
                self.partial_table = {"key": table.key, "offset": rows_scanned}
                self.scan_stats[table.key] = {
                    "mode": "directed_probe", "rows_scanned": rows_scanned,
                    "candidate_columns": [column.name for column in selected], "matches": matches, "engine": table.engine,
                }
                self.save_progress()
                self.check_deadline()
        self.expand_completed_table(table, selected)
        self.scanned_tables.add(table.key)
        self.partial_table = None
        self.scan_stats[table.key] = {
            "mode": "directed_probe", "rows_scanned": rows_scanned,
            "candidate_columns": [column.name for column in selected], "matches": matches, "engine": table.engine,
        }
        self.save_progress()

    def add_profile_relations(self) -> None:
        entries: list[tuple[TableMeta, ColumnMeta, BottomK]] = []
        for table in self.tables.values():
            for column in table.columns:
                profile = self.profiles.get(f"{table.key}\0{column.name}")
                if profile:
                    entries.append((table, column, profile))
        for (left_table, left_column, left_profile), (right_table, right_column, right_profile) in itertools.combinations(entries, 2):
            if left_table.file_path == right_table.file_path:
                continue
            synthetic = ValueSource("", "", "", left_column.name, left_column.kind, left_column.base, "", "", 0, "")
            compatibility = column_compatibility(synthetic, right_column)
            if compatibility < 0.72:
                continue
            shared = sorted(set(left_profile.items) & set(right_profile.items))
            if not shared:
                continue
            for value_hash in shared[:MAX_EVIDENCE]:
                left = left_profile.items[value_hash]
                right = right_profile.items[value_hash]
                minimum_specificity = min(left.specificity, right.specificity)
                confirmed = (
                    compatibility >= 0.95 and len(shared) >= 2 and minimum_specificity >= 2
                ) or (
                    compatibility >= 0.82 and len(shared) >= 3 and minimum_specificity >= 3
                )
                confidence = min(0.97, 0.78 + 0.12 * compatibility + 0.02 * min(len(shared), 4))
                self.relations.add(
                    left, right, "profiled_value_overlap", confidence,
                    "confirmed" if confirmed else "supported_hypothesis", "bottom_k_exact_overlap",
                    f"Bounded column profiles contain {len(shared)} exact shared fingerprint(s) (column compatibility {compatibility:.2f}).",
                )

    def run(self) -> dict[str, Any]:
        if not self.input_root.is_dir():
            raise ValueError(f"Input directory does not exist: {self.input_root}")
        if self.input_root == self.output_root or self.input_root in self.output_root.parents:
            raise ValueError("Output directory must be outside the input data directory")
        goal_text = ""
        goal_path = Path(self.args.goal_file) if self.args.goal_file else None
        if goal_path and goal_path.is_file():
            goal_text = goal_path.read_text(encoding="utf-8", errors="replace")[:100_000]
        self.output_root.mkdir(parents=True, exist_ok=True)
        self.load_or_initialize(goal_tokens(goal_text))
        status = "complete"
        try:
            if not self.resumed:
                seed_tables, seed_documents = self.seed_candidates()
                all_names = [item.path for item in self.files]
                for table in seed_tables:
                    source_file = self.file_by_id[table.file_id]
                    is_bootstrap = (
                        table.estimated_cells is not None
                        and table.estimated_cells > self.args.seed_cell_budget
                        and source_file.size > self.args.seed_file_bytes
                    )
                    self.process_seed_table(table, bootstrap=is_bootstrap)
                    self.check_deadline()
                for item in seed_documents:
                    self.process_seed_document(item, all_names)
                    self.check_deadline()
            remaining = [table for table in self.tables.values() if table.key not in self.scanned_tables]
            if self.partial_table and self.partial_table.get("key") in self.tables:
                table = self.tables[self.partial_table["key"]]
                self.scan_large_table(table, int(self.partial_table.get("offset", 0)))
                remaining = [item for item in remaining if item.key != table.key]
            while remaining:
                self.check_deadline()
                table = max(remaining, key=self.table_score)
                remaining.remove(table)
                self.scan_large_table(table)
            self.add_profile_relations()
            self.save_progress()
        except DeadlineReached:
            status = "partial"
            self.warnings.append("Time budget reached; rerun the same command to resume from progress.json.")
            self.save_progress()
        return self.write_result(status)

    def write_result(self, status: str) -> dict[str, Any]:
        relations = self.relations.values()
        chains = build_chains(relations)
        files_completed = {self.tables[key].file_path for key in self.scanned_tables if key in self.tables}
        result = {
            "schema_version": SCHEMA_VERSION,
            "status": status,
            "generated_at": utc_now(),
            "input_root": str(self.input_root),
            "strategy": "progressive_seeded_probe",
            "coverage": {
                "files_discovered": len(self.files),
                "tables_discovered": len(self.tables),
                "tables_completed": len([key for key in self.scanned_tables if key in self.tables]),
            "files_with_completed_table_scan": len(files_completed),
                "seed_value_count": len(self.seeds.by_fingerprint),
                "profile_value_count": sum(len(profile.items) for profile in self.profiles.values()),
                "frontier_value_count": sum(len(profile.items) for profile in self.frontier_profiles.values()),
                "warnings": list(dict.fromkeys(self.warnings)),
                "partial_table": self.partial_table,
                "guarantee": "All emitted relations have exact fingerprint evidence; unseeded values outside bounded profiles are not claimed as covered.",
            },
            "files": [self.file_json(item) for item in self.files],
            "scan_stats": self.scan_stats,
            "relations": relations,
            "chains": chains,
            "artifacts": {
                "catalog": str(self.output_root / "catalog.json"),
                "json": str(self.output_root / "relations.json"),
                "markdown": str(self.output_root / "relation-report.md"),
                "mermaid": str(self.output_root / "relations.mmd"),
                "evidence_index": str(self.output_root / "evidence.sqlite3"),
                "progress": str(self.output_root / "progress.json"),
            },
        }
        atomic_json(self.output_root / "relations.json", result)
        write_markdown(result, self.output_root / "relation-report.md")
        write_mermaid(result, self.output_root / "relations.mmd")
        write_evidence_database(result, self.output_root / "evidence.sqlite3")
        return result


class UnionFind:
    def __init__(self, values: Iterable[str]) -> None:
        self.parent = {value: value for value in values}

    def find(self, value: str) -> str:
        while self.parent[value] != value:
            self.parent[value] = self.parent[self.parent[value]]
            value = self.parent[value]
        return value

    def union(self, left: str, right: str) -> bool:
        left_root, right_root = self.find(left), self.find(right)
        if left_root == right_root:
            return False
        self.parent[right_root] = left_root
        return True


def build_chains(relations: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    nodes = sorted({item["source"] for item in relations} | {item["target"] for item in relations})
    union = UnionFind(nodes)
    for relation in relations:
        union.union(relation["source"], relation["target"])
    groups: dict[str, set[str]] = defaultdict(set)
    for node in nodes:
        groups[union.find(node)].add(node)
    chains: list[dict[str, Any]] = []
    for files in groups.values():
        edges = [item for item in relations if item["source"] in files and item["target"] in files]
        tree = UnionFind(files)
        core: list[str] = []
        for edge in sorted(edges, key=lambda item: (-item["confidence"], item["id"])):
            if tree.union(edge["source"], edge["target"]):
                core.append(edge["id"])
        signature = "\0".join(sorted(files))
        chains.append({
            "id": "C-" + hashlib.sha1(signature.encode("utf-8")).hexdigest()[:12],
            "files": sorted(files),
            "relation_ids": [item["id"] for item in edges],
            "core_relation_ids": core,
            "relation_count": len(edges),
            "explanation": f"Connected evidence network with {len(files)} files and {len(edges)} direct relationship(s).",
        })
    return sorted(chains, key=lambda item: item["id"])


def write_evidence_database(result: dict[str, Any], path: Path) -> None:
    for suffix in ("", "-wal", "-shm"):
        target = Path(str(path) + suffix)
        if target.exists():
            target.unlink()
    connection = sqlite3.connect(path)
    try:
        connection.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE relations(id TEXT PRIMARY KEY, source TEXT, target TEXT, type TEXT, verdict TEXT, confidence REAL, source_column TEXT, target_column TEXT, evidence_count INTEGER, explanation TEXT);
            CREATE TABLE evidence(relation_id TEXT, kind TEXT, source_locator TEXT, target_locator TEXT, value_preview TEXT, fingerprint TEXT, strength TEXT, note TEXT);
            """
        )
        for relation in result["relations"]:
            connection.execute(
                "INSERT INTO relations VALUES (?,?,?,?,?,?,?,?,?,?)",
                (relation["id"], relation["source"], relation["target"], relation["type"], relation["verdict"], relation["confidence"], relation.get("source_column", ""), relation.get("target_column", ""), relation["evidence_count"], relation["explanation"]),
            )
            for evidence in relation["evidence"]:
                connection.execute(
                    "INSERT INTO evidence VALUES (?,?,?,?,?,?,?,?)",
                    (relation["id"], evidence["kind"], evidence["source_locator"], evidence["target_locator"], evidence["value_preview"], evidence["fingerprint"], evidence["strength"], evidence.get("note", "")),
                )
        connection.commit()
    finally:
        connection.close()


def write_markdown(result: dict[str, Any], path: Path) -> None:
    coverage = result["coverage"]
    lines = [
        "# 数据关系链路分析报告", "",
        f"- 状态：`{result['status']}`", f"- 策略：`{result['strategy']}`",
        f"- 文件：{coverage['files_discovered']}；表：{coverage['tables_completed']}/{coverage['tables_discovered']}",
        f"- 有界种子：{coverage['seed_value_count']}；列摘要值：{coverage['profile_value_count']}",
        f"- 关系：{len(result['relations'])}；链路：{len(result['chains'])}", "",
        "## 覆盖边界", "", coverage["guarantee"], "",
    ]
    if coverage["warnings"]:
        lines.extend(["## 警告", ""] + [f"- {warning}" for warning in coverage["warnings"]] + [""])
    lines.extend(["## 关系", ""])
    if not result["relations"]:
        lines.append("未发现达到证据门槛的关系。")
    for relation in result["relations"]:
        lines.extend([
            f"### {relation['id']} · {relation['type']}", "",
            f"`{relation['source']}` ↔ `{relation['target']}`", "",
            f"- 列：`{relation.get('source_column', '')}` ↔ `{relation.get('target_column', '')}`",
            f"- 判定：`{relation['verdict']}`；置信度：`{relation['confidence']:.3f}`",
            f"- 说明：{relation['explanation']}", f"- 证据总数：{relation['evidence_count']}",
        ])
        for evidence in relation["evidence"]:
            lines.append(f"- 证据：`{evidence['kind']}`；`{evidence['source_locator']}` ↔ `{evidence['target_locator']}`；匹配 `{evidence['value_preview']}`；指纹 `{evidence['fingerprint'][:16]}...`")
        if relation["omitted_evidence_count"]:
            lines.append(f"- 未展开证据：{relation['omitted_evidence_count']}")
        lines.append("")
    lines.extend(["## 链路", ""])
    for chain in result["chains"]:
        lines.extend([f"### {chain['id']}", "", chain["explanation"], "", "文件：" + "、".join(f"`{item}`" for item in chain["files"]), "", "核心关系：" + "、".join(f"`{item}`" for item in chain["core_relation_ids"]), ""])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_mermaid(result: dict[str, Any], path: Path) -> None:
    names = sorted({item["source"] for item in result["relations"]} | {item["target"] for item in result["relations"]})
    ids = {name: f"F{index}" for index, name in enumerate(names, 1)}
    lines = ["flowchart LR"]
    for name in names:
        label = name.replace('"', "'")
        lines.append(f'  {ids[name]}["{label[-60:]}"]')
    for relation in result["relations"]:
        label = f"{relation['type']} {relation['confidence']:.2f}".replace('"', "'")
        lines.append(f'  {ids[relation["source"]]} ---|"{label}"| {ids[relation["target"]]}')
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def compact_summary(result: dict[str, Any], offset: int, limit: int) -> dict[str, Any]:
    relations = result.get("relations", [])
    warnings = result.get("coverage", {}).get("warnings", [])
    coverage = dict(result.get("coverage", {}))
    coverage["warning_count"] = len(warnings)
    coverage["warnings"] = warnings[:20]
    coverage["omitted_warning_count"] = max(0, len(warnings) - 20)
    return {
        "status": result.get("status", "complete"),
        "coverage": coverage,
        "relation_count": len(relations),
        "chain_count": len(result.get("chains", [])),
        "relation_page": {"offset": offset, "limit": limit, "items": [
            {key: item.get(key) for key in ("id", "source", "target", "type", "verdict", "confidence", "source_column", "target_column", "explanation", "evidence_count")}
            for item in relations[offset:offset + limit]
        ], "has_more": offset + limit < len(relations)},
        "artifacts": result.get("artifacts", {}),
    }


def load_result(path: Path) -> dict[str, Any]:
    result = json.loads(path.read_text(encoding="utf-8"))
    if result.get("schema_version") != SCHEMA_VERSION:
        raise ValueError("Unsupported result schema")
    return result


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Progressively discover evidence-backed relationships")
    commands = parser.add_subparsers(dest="command", required=True)
    analyze = commands.add_parser("analyze")
    analyze.add_argument("--input", default="/workspace/data")
    analyze.add_argument("--output", default="/workspace/outputs/data-relations")
    analyze.add_argument("--goal-file", default="/workspace/description.md")
    analyze.add_argument("--ocr-mode", choices=["auto", "always", "never"], default="auto")
    analyze.add_argument("--deadline-seconds", type=int, default=780)
    analyze.add_argument("--resume", action=argparse.BooleanOptionalAction, default=True)
    analyze.add_argument("--seed-cell-budget", type=int, default=100_000)
    analyze.add_argument("--seed-file-bytes", type=int, default=8 * 1024 * 1024)
    analyze.add_argument("--seed-values-per-column", type=int, default=256)
    analyze.add_argument("--max-seed-values", type=int, default=20_000)
    analyze.add_argument("--max-text-seeds", type=int, default=2_000)
    analyze.add_argument("--profile-size", type=int, default=128)
    analyze.add_argument("--frontier-values-per-column", type=int, default=128)
    analyze.add_argument("--min-expansion-distinct-ratio", type=float, default=0.01)
    analyze.add_argument("--max-matches-per-seed", type=int, default=50)
    analyze.add_argument("--bootstrap-rows", type=int, default=1_000)
    analyze.add_argument("--checkpoint-rows", type=int, default=50_000)
    analyze.add_argument("--document-character-budget", type=int, default=5_000_000)
    analyze.add_argument("--summary-limit", type=int, default=20)
    summary = commands.add_parser("summary")
    summary.add_argument("--result", default="/workspace/outputs/data-relations/relations.json")
    summary.add_argument("--offset", type=int, default=0)
    summary.add_argument("--limit", type=int, default=20)
    relation = commands.add_parser("relation")
    relation.add_argument("relation_id")
    relation.add_argument("--result", default="/workspace/outputs/data-relations/relations.json")
    chain = commands.add_parser("chain")
    chain.add_argument("chain_id")
    chain.add_argument("--result", default="/workspace/outputs/data-relations/relations.json")
    chain.add_argument("--offset", type=int, default=0)
    chain.add_argument("--limit", type=int, default=50)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8")
    args = build_parser().parse_args(argv)
    try:
        if args.command == "analyze":
            result = ProgressiveAnalyzer(Path(args.input), Path(args.output), args).run()
            print(json.dumps(compact_summary(result, 0, max(0, min(args.summary_limit, 100))), ensure_ascii=False, indent=2))
            return 0
        result = load_result(Path(args.result))
        if args.command == "summary":
            print(json.dumps(compact_summary(result, max(0, args.offset), max(1, min(args.limit, 100))), ensure_ascii=False, indent=2))
            return 0
        collection = "relations" if args.command == "relation" else "chains"
        identifier = args.relation_id if args.command == "relation" else args.chain_id
        item = next((entry for entry in result.get(collection, []) if entry.get("id") == identifier), None)
        if item is None:
            print(json.dumps({"status": "not_found", "id": identifier}, ensure_ascii=False), file=sys.stderr)
            return 2
        if args.command == "chain":
            offset, limit = max(0, args.offset), max(1, min(args.limit, 100))
            item = {**item, "files": item["files"][offset:offset + limit], "relation_ids": item["relation_ids"][offset:offset + limit], "core_relation_ids": item["core_relation_ids"][offset:offset + limit], "offset": offset, "limit": limit}
        print(json.dumps({"status": "success", collection[:-1]: item}, ensure_ascii=False, indent=2))
        return 0
    except Exception as exc:
        print(json.dumps({"status": "error", "message": f"{type(exc).__name__}: {exc}"}, ensure_ascii=False), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
