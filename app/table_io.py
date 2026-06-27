"""表数据读取工具。

核心约束（来自需求 Note 2）：历史数据动辄上万行，**绝不**整表加载逐行查看。
本模块只读取表头 + 少量随机样本，并以「轻量方式」估算行数。

另一个工程现实：真实表格的表头**不一定在第一行**——上方常有标题、说明、空行、
合并单元格等无关内容。因此本模块在读取前会**自动识别表头所在行**（`_detect_header_row`），
让后续所有能力（抽样、行数、关联推导、整表查询）都基于「真正的表头」工作。
"""

from __future__ import annotations

import csv
import functools
import random
from pathlib import Path
from typing import Any

import pandas as pd

from .models import ColumnMeta, TableMeta

# 为估算行数 / 抽样而最多读取的行数上限：足够代表分布，又不至于拖垮内存
_SCAN_CAP = 2000
# 随机抽样条数
_SAMPLE_N = 3
# 表头识别时最多探查的行数（标题区一般不会很长）
_DETECT_SCAN = 20
_EXCEL_SUFFIX = {".xlsx", ".xls"}


# ===========================================================================
# 编码与原始网格读取
# ===========================================================================
def _csv_encoding(path: Path) -> str:
    """嗅探 CSV/TSV 文本编码（兼容 UTF-8 BOM 与 GBK 中文表格）。"""
    head = path.read_bytes()[:8192]
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            head.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "utf-8"


def _csv_sep(path: Path) -> str:
    return "\t" if path.suffix.lower() == ".tsv" else ","


def _read_detection_grid(path: Path) -> list[list[Any]]:
    """读取前若干行的「原始二维网格」（不指定表头），用于识别表头位置。"""
    if path.suffix.lower() in _EXCEL_SUFFIX:
        raw = pd.read_excel(path, header=None, nrows=_DETECT_SCAN)
        return [list(raw.iloc[i]) for i in range(len(raw))]
    rows: list[list[Any]] = []
    with path.open("r", encoding=_csv_encoding(path), newline="", errors="replace") as fp:
        reader = csv.reader(fp, delimiter=_csv_sep(path))
        for i, row in enumerate(reader):
            if i >= _DETECT_SCAN:
                break
            rows.append(row)
    return rows


# ===========================================================================
# 表头识别
# ===========================================================================
def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    return str(value).strip() == ""


def _looks_numeric(value: Any) -> bool:
    if _is_blank(value):
        return False
    text = str(value).strip().replace(",", "").replace("%", "")
    try:
        float(text)
        return True
    except ValueError:
        return False


def _table_width(grid: list[list[Any]]) -> int:
    """估算表格的真实列数：取各行宽度的众数。

    关键：CSV 的标题行往往只有一个单元格（无分隔符），若用「行自身宽度」做分母，
    其填充度会被误判为 100%。改用整表列数做分母，标题行的填充度才会真实偏低。
    """
    from collections import Counter

    lengths = [len(r) for r in grid if any(not _is_blank(c) for c in r)]
    if not lengths:
        return 1
    return Counter(lengths).most_common(1)[0][0]


def _score_header_row(grid: list[list[Any]], idx: int, table_ncols: int) -> float:
    """给某一行「作为表头」的可能性打分。

    表头行的典型特征：单元格基本填满（相对整表列数）、以文本为主、取值互不相同；
    其下方还应紧跟着像数据的行。标题行通常只占一格，相对整表列数填充度很低，故得分低。
    """
    row = grid[idx]
    cells = [c for c in row if not _is_blank(c)]
    if not cells:
        return -1.0
    width = max(table_ncols, 1)

    fill_ratio = min(len(cells) / width, 1.0)                                 # 填充度（相对整表列数）
    text_ratio = sum(1 for c in cells if not _looks_numeric(c)) / len(cells)  # 文本占比
    unique_ratio = len({str(c).strip() for c in cells}) / len(cells)          # 唯一度

    # 下方数据行的填充度（表头之下应是成片的数据，同样相对整表列数归一）
    below = grid[idx + 1: idx + 4]
    if below:
        below_fill = sum(
            sum(0 if _is_blank(c) else 1 for c in r) / width for r in below
        ) / len(below)
    else:
        below_fill = 0.0

    score = (
        0.40 * fill_ratio
        + 0.30 * text_ratio
        + 0.20 * unique_ratio
        + 0.10 * min(below_fill, 1.0)
    )
    # 同分时偏向更靠前的行，避免误把某一行数据当作表头
    return score - idx * 0.01


def _detect_header_row(path: Path) -> int:
    """识别表头所在行号（0 表示首行即表头）。失败时安全回退为 0。"""
    try:
        grid = _read_detection_grid(path)
    except Exception:  # noqa: BLE001
        return 0
    if not grid:
        return 0
    table_ncols = _table_width(grid)
    best_idx, best_score = 0, -2.0
    for i in range(len(grid)):
        score = _score_header_row(grid, i, table_ncols)
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx


@functools.lru_cache(maxsize=256)
def _header_cache(path_str: str, mtime: float) -> int:
    return _detect_header_row(Path(path_str))


def resolve_header_row(file_path: str | Path) -> int:
    """获取表头行号（按文件路径 + 修改时间缓存，避免重复探查）。"""
    path = Path(file_path)
    try:
        mtime = path.stat().st_mtime
    except OSError:
        mtime = 0.0
    return _header_cache(str(path), mtime)


# ===========================================================================
# 数据帧读取（自动跳过表头上方的标题/空行）
# ===========================================================================
def _load_scan_frame(path: Path, header_row: int | None = None) -> pd.DataFrame:
    """读取用于结构分析的「采样数据帧」，最多 `_SCAN_CAP` 行，自动定位表头。"""
    hr = resolve_header_row(path) if header_row is None else header_row
    if path.suffix.lower() in _EXCEL_SUFFIX:
        return pd.read_excel(path, skiprows=hr or None, nrows=_SCAN_CAP)
    return pd.read_csv(
        path,
        skiprows=hr or None,
        nrows=_SCAN_CAP,
        sep=_csv_sep(path),
        encoding=_csv_encoding(path),
    )


def load_full_frame(file_path: str) -> pd.DataFrame:
    """加载整表（仅供「执行业务/查询数据」这类确需全量计算的场景使用），自动定位表头。"""
    path = Path(file_path)
    hr = resolve_header_row(path)
    if path.suffix.lower() in _EXCEL_SUFFIX:
        return pd.read_excel(path, skiprows=hr or None)
    return pd.read_csv(
        path, skiprows=hr or None, sep=_csv_sep(path), encoding=_csv_encoding(path)
    )


# ===========================================================================
# 行数估算
# ===========================================================================
def _read_excel_dimensions(path: Path) -> int:
    """用 openpyxl 只读模式快速获取 Excel 总行数（不加载单元格值）。"""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    try:
        return ws.max_row if (ws := wb.active) and ws.max_row else 0
    finally:
        wb.close()


def _count_physical_lines(path: Path) -> int:
    """按缓冲方式统计文本文件物理行数（不解析内容，内存恒定）。"""
    total = 0
    with path.open("rb") as fp:
        for _ in fp:
            total += 1
    return total


def _estimate_row_count(path: Path, header_row: int) -> int:
    """估算「数据行数」= 总行数 - 表头上方的标题行 - 表头行本身。"""
    try:
        if path.suffix.lower() in _EXCEL_SUFFIX:
            total = _read_excel_dimensions(path)
        else:
            total = _count_physical_lines(path)
    except Exception:  # noqa: BLE001  估算失败不应阻断流程
        return 0
    return max(total - header_row - 1, 0)


# ===========================================================================
# 工具函数
# ===========================================================================
def _jsonable(value: Any) -> Any:
    """将 pandas/numpy 标量转换为可 JSON 序列化的原生类型。"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if pd.api.types.is_scalar(value) and pd.isna(value):
        return None
    if hasattr(value, "item"):  # numpy 标量
        try:
            return value.item()
        except Exception:  # noqa: BLE001
            return str(value)
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


# ===========================================================================
# 对外主入口
# ===========================================================================
def inspect_table(file_path: str, table_name: str | None = None) -> TableMeta:
    """读取单张表的结构元信息：表头、字段类型、空值率、随机样本。

    会先自动识别表头所在行（跳过上方标题/空行），再仅扫描前 `_SCAN_CAP` 行用于分析，
    并随机抽取 `_SAMPLE_N` 条样本，严格遵守「不遍历全量数据」的约束。
    """
    path = Path(file_path)
    name = table_name or path.stem

    header_row = resolve_header_row(path)
    frame = _load_scan_frame(path, header_row)
    row_count = _estimate_row_count(path, header_row)

    columns: list[ColumnMeta] = []
    for col in frame.columns:
        series = frame[col]
        non_null = series.dropna()
        # 在去重后的非空值中抽样，更能代表字段取值空间
        unique_vals = list(dict.fromkeys(non_null.tolist()))
        sample_pool = unique_vals if unique_vals else non_null.tolist()
        sample = random.sample(sample_pool, min(_SAMPLE_N, len(sample_pool))) if sample_pool else []
        null_rate = float(series.isna().mean()) if len(series) else 0.0
        columns.append(
            ColumnMeta(
                name=str(col),
                dtype=str(series.dtype),
                null_rate=round(null_rate, 4),
                sample_values=[_jsonable(v) for v in sample],
            )
        )

    # 随机抽取若干完整样本行
    sample_rows: list[dict[str, Any]] = []
    if len(frame):
        idxs = random.sample(range(len(frame)), min(_SAMPLE_N, len(frame)))
        for i in idxs:
            row = {str(k): _jsonable(v) for k, v in frame.iloc[i].to_dict().items()}
            sample_rows.append(row)

    return TableMeta(
        table_name=name,
        display_name=path.name,
        file_path=str(path),
        row_count=row_count,
        col_count=len(frame.columns),
        header_row=header_row,
        columns=columns,
        sample_rows=sample_rows,
    )


def column_value_set(file_path: str, column: str, limit: int = _SCAN_CAP) -> set[Any]:
    """读取某列的取值集合（限量），用于关联关系的「样本值重叠率」判断。"""
    frame = _load_scan_frame(Path(file_path))
    if column not in frame.columns:
        return set()
    return {_jsonable(v) for v in frame[column].dropna().head(limit).tolist()}
