"""Bounded previews for files anywhere inside a Studio workspace."""

from __future__ import annotations

import csv
import json
import mimetypes
import sqlite3
import tarfile
import zipfile
from itertools import islice
from pathlib import Path
from typing import Any, Iterator
from xml.etree import ElementTree


TEXT_PREVIEW_BYTES = 2 * 1024 * 1024
PREVIEW_ROW_LIMIT = 20
PREVIEW_COLUMN_LIMIT = 80
ARCHIVE_ENTRY_LIMIT = 200
PREVIEW_SHEET_LIMIT = 10

MARKDOWN_EXTENSIONS = {".md", ".markdown"}
MERMAID_EXTENSIONS = {".mmd", ".mermaid"}
JSON_EXTENSIONS = {".json", ".geojson"}
CSV_EXTENSIONS = {".csv", ".tsv"}
SPREADSHEET_EXTENSIONS = {".xlsx", ".xls"}
SQLITE_EXTENSIONS = {".sqlite", ".sqlite3", ".db"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tif", ".tiff", ".svg", ".ico"}
VIDEO_EXTENSIONS = {".mp4", ".webm", ".ogv", ".mov", ".m4v"}
AUDIO_EXTENSIONS = {".mp3", ".wav", ".ogg", ".oga", ".m4a", ".flac", ".aac"}
DOCUMENT_EXTENSIONS = {".docx", ".pptx"}
ARCHIVE_EXTENSIONS = {".zip", ".tar", ".tgz", ".gz", ".bz2", ".xz"}
SOURCE_EXTENSIONS = {
    ".txt", ".log", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".conf",
    ".xml", ".html", ".htm", ".css", ".scss", ".sass", ".less", ".js",
    ".mjs", ".cjs", ".ts", ".tsx", ".jsx", ".vue", ".py", ".pyi", ".java",
    ".kt", ".kts", ".go", ".rs", ".c", ".h", ".cpp", ".hpp", ".cs", ".php",
    ".rb", ".sh", ".ps1", ".bat", ".cmd", ".sql", ".graphql", ".gql", ".env",
    ".properties", ".jsonl", ".ndjson",
}


def preview_workspace_file(path: Path) -> dict[str, Any]:
    """Return a bounded, renderer-neutral preview payload."""

    suffix = path.suffix.casefold()
    mime_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    payload: dict[str, Any] = {
        "filename": path.name,
        "extension": suffix,
        "size": path.stat().st_size,
        "mime_type": mime_type,
        "kind": "unsupported",
        "text": "",
        "columns": [],
        "sample_rows": [],
        "sheets": [],
        "truncated": False,
        "warnings": [],
    }

    try:
        if suffix in MERMAID_EXTENSIONS:
            return _text_payload(path, payload, "mermaid")
        if suffix in MARKDOWN_EXTENSIONS:
            return _text_payload(path, payload, "markdown")
        if suffix in JSON_EXTENSIONS:
            return _json_payload(path, payload)
        if suffix in CSV_EXTENSIONS:
            return _delimited_payload(path, payload, "\t" if suffix == ".tsv" else ",")
        if suffix in SPREADSHEET_EXTENSIONS:
            return _spreadsheet_payload(path, payload)
        if suffix == ".parquet":
            return _parquet_payload(path, payload)
        if suffix in SQLITE_EXTENSIONS:
            return _sqlite_payload(path, payload)
        if suffix == ".docx":
            return _docx_payload(path, payload)
        if suffix == ".pptx":
            return _pptx_payload(path, payload)
        if suffix == ".pdf":
            payload["kind"] = "pdf"
            return payload
        if suffix in IMAGE_EXTENSIONS:
            payload["kind"] = "image"
            return payload
        if suffix in VIDEO_EXTENSIONS:
            payload["kind"] = "video"
            return payload
        if suffix in AUDIO_EXTENSIONS:
            payload["kind"] = "audio"
            return payload
        if suffix in ARCHIVE_EXTENSIONS or tarfile.is_tarfile(path):
            return _archive_payload(path, payload)
        if suffix in SOURCE_EXTENSIONS or mime_type.startswith("text/"):
            return _text_payload(path, payload, "text")
    except Exception as exc:
        payload["kind"] = "error"
        payload["warnings"].append(f"{type(exc).__name__}: {exc}")
    return payload


def _read_text_prefix(path: Path, limit: int = TEXT_PREVIEW_BYTES) -> tuple[str, bool, str]:
    with path.open("rb") as stream:
        sample = stream.read(limit + 1)
    truncated = len(sample) > limit
    sample = sample[:limit]
    if sample.startswith(b"\xef\xbb\xbf"):
        encoding = "utf-8-sig"
    elif sample.startswith((b"\xff\xfe", b"\xfe\xff")):
        encoding = "utf-16"
    else:
        encoding = "utf-8"
        for candidate in ("utf-8", "gb18030"):
            try:
                sample.decode(candidate)
                encoding = candidate
                break
            except UnicodeDecodeError:
                continue
    return sample.decode(encoding, errors="replace"), truncated, encoding


def _text_payload(path: Path, payload: dict[str, Any], kind: str) -> dict[str, Any]:
    text, truncated, encoding = _read_text_prefix(path)
    payload.update(kind=kind, text=text, truncated=truncated, encoding=encoding)
    if truncated:
        payload["warnings"].append(f"Preview is limited to the first {TEXT_PREVIEW_BYTES // (1024 * 1024)} MB.")
    return payload


def _json_payload(path: Path, payload: dict[str, Any]) -> dict[str, Any]:
    payload = _text_payload(path, payload, "json")
    if not payload["truncated"]:
        try:
            payload["text"] = json.dumps(json.loads(payload["text"]), ensure_ascii=False, indent=2)
        except (TypeError, ValueError, json.JSONDecodeError) as exc:
            payload["warnings"].append(f"JSON formatting failed: {exc}")
    return payload


def _unique_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(values[:PREVIEW_COLUMN_LIMIT], 1):
        base = str(value or "").strip() or f"column_{index}"
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return headers


def _rows_to_dicts(headers: list[str], rows: Iterator[list[Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for values in rows:
        result.append({header: _json_value(values[index] if index < len(values) else None) for index, header in enumerate(headers)})
        if len(result) >= PREVIEW_ROW_LIMIT:
            break
    return result


def _infer_header_and_rows(rows: list[list[Any]]) -> tuple[list[str], list[dict[str, Any]], int]:
    if not rows:
        return [], [], 0
    candidate_count = min(10, len(rows))

    def score(values: list[Any]) -> tuple[int, int, int]:
        nonempty = [str(value).strip() for value in values[:PREVIEW_COLUMN_LIMIT] if value not in (None, "")]
        unique = len(set(nonempty))
        textual = sum(not value.replace(".", "", 1).isdigit() for value in nonempty)
        return len(nonempty), unique, textual

    header_index = max(range(candidate_count), key=lambda index: (score(rows[index]), -index))
    headers = _unique_headers(rows[header_index])
    preview_rows = _rows_to_dicts(headers, iter(rows[header_index + 1:]))
    return headers, preview_rows, header_index + 1


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _delimited_payload(path: Path, payload: dict[str, Any], delimiter: str) -> dict[str, Any]:
    _text, _truncated, encoding = _read_text_prefix(path, 65_536)
    with path.open("r", encoding=encoding, errors="replace", newline="") as stream:
        reader = csv.reader(stream, delimiter=delimiter)
        headers = _unique_headers(next(reader, []))
        rows = _rows_to_dicts(headers, reader)
    payload.update(kind="table", columns=headers, sample_rows=rows, truncated=path.stat().st_size > 65_536)
    if payload["truncated"] or len(rows) >= PREVIEW_ROW_LIMIT:
        payload["warnings"].append(f"Table preview shows at most {PREVIEW_ROW_LIMIT} rows.")
    return payload


def _spreadsheet_payload(path: Path, payload: dict[str, Any]) -> dict[str, Any]:
    suffix = path.suffix.casefold()
    sheets: list[dict[str, Any]] = []
    total_sheets = 0
    if suffix == ".xlsx":
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("XLSX preview requires openpyxl") from exc
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            total_sheets = len(workbook.sheetnames)
            for sheet_name in workbook.sheetnames[:PREVIEW_SHEET_LIMIT]:
                worksheet = workbook[sheet_name]
                iterator = worksheet.iter_rows(values_only=True)
                buffered = [list(row) for row in islice(iterator, PREVIEW_ROW_LIMIT + 10)]
                headers, rows, header_row = _infer_header_and_rows(buffered)
                sheets.append({
                    "name": worksheet.title,
                    "columns": headers,
                    "sample_rows": rows,
                    "header_row": header_row,
                    "row_count": worksheet.max_row,
                    "column_count": worksheet.max_column,
                })
        finally:
            workbook.close()
    else:
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("XLS preview requires xlrd") from exc
        workbook = xlrd.open_workbook(path, on_demand=True)
        try:
            total_sheets = workbook.nsheets
            for sheet_index in range(min(workbook.nsheets, PREVIEW_SHEET_LIMIT)):
                worksheet = workbook.sheet_by_index(sheet_index)
                buffered = [
                    worksheet.row_values(index)
                    for index in range(min(worksheet.nrows, PREVIEW_ROW_LIMIT + 10))
                ]
                headers, rows, header_row = _infer_header_and_rows(buffered)
                sheets.append({
                    "name": worksheet.name,
                    "columns": headers,
                    "sample_rows": rows,
                    "header_row": header_row,
                    "row_count": worksheet.nrows,
                    "column_count": worksheet.ncols,
                })
        finally:
            workbook.release_resources()
    first = sheets[0] if sheets else {"columns": [], "sample_rows": []}
    payload.update(
        kind="table",
        columns=first["columns"],
        sample_rows=first["sample_rows"],
        sheets=sheets,
        truncated=any(int(sheet.get("row_count", 0)) > PREVIEW_ROW_LIMIT + 1 for sheet in sheets),
    )
    if total_sheets > PREVIEW_SHEET_LIMIT:
        payload["warnings"].append(f"Workbook preview shows the first {PREVIEW_SHEET_LIMIT} worksheets.")
    if payload["truncated"]:
        payload["warnings"].append(f"Each worksheet preview shows at most {PREVIEW_ROW_LIMIT} rows.")
    return payload


def _parquet_payload(path: Path, payload: dict[str, Any]) -> dict[str, Any]:
    try:
        import duckdb
    except ImportError as exc:
        raise RuntimeError("Parquet preview requires duckdb") from exc
    connection = duckdb.connect()
    try:
        cursor = connection.execute("SELECT * FROM read_parquet(?) LIMIT ?", [str(path), PREVIEW_ROW_LIMIT])
        columns = [str(item[0]) for item in cursor.description[:PREVIEW_COLUMN_LIMIT]]
        rows = [
            {column: _json_value(row[index]) for index, column in enumerate(columns)}
            for row in cursor.fetchall()
        ]
    finally:
        connection.close()
    payload.update(kind="table", columns=columns, sample_rows=rows, truncated=len(rows) >= PREVIEW_ROW_LIMIT)
    if payload["truncated"]:
        payload["warnings"].append(f"Parquet preview shows at most {PREVIEW_ROW_LIMIT} rows.")
    return payload


def _sqlite_payload(path: Path, payload: dict[str, Any]) -> dict[str, Any]:
    connection = sqlite3.connect(path.resolve().as_uri() + "?mode=ro", uri=True)
    try:
        table_names = [
            str(row[0])
            for row in connection.execute(
                f"SELECT name FROM sqlite_master WHERE type IN ('table', 'view') AND name NOT LIKE 'sqlite_%' ORDER BY name LIMIT {PREVIEW_SHEET_LIMIT}"
            )
        ]
        sheets: list[dict[str, Any]] = []
        for table_name in table_names:
            quoted = '"' + table_name.replace('"', '""') + '"'
            cursor = connection.execute(f"SELECT * FROM {quoted} LIMIT ?", (PREVIEW_ROW_LIMIT,))
            columns = [str(item[0]) for item in (cursor.description or [])[:PREVIEW_COLUMN_LIMIT]]
            rows = [
                {column: _json_value(row[index]) for index, column in enumerate(columns)}
                for row in cursor.fetchall()
            ]
            sheets.append({"name": table_name, "columns": columns, "sample_rows": rows})
    finally:
        connection.close()
    first = sheets[0] if sheets else {"columns": [], "sample_rows": []}
    payload.update(kind="database", columns=first["columns"], sample_rows=first["sample_rows"], sheets=sheets, truncated=True)
    payload["warnings"].append(f"Database preview shows at most {PREVIEW_ROW_LIMIT} rows per table.")
    return payload


def _docx_payload(path: Path, payload: dict[str, Any]) -> dict[str, Any]:
    paragraphs: list[str] = []
    with zipfile.ZipFile(path) as archive, archive.open("word/document.xml") as stream:
        for _event, element in ElementTree.iterparse(stream, events=("end",)):
            if element.tag.endswith("}p"):
                text = "".join(node.text or "" for node in element.iter() if node.tag.endswith("}t")).strip()
                if text:
                    paragraphs.append(text)
                element.clear()
                if sum(map(len, paragraphs)) >= TEXT_PREVIEW_BYTES:
                    break
    text = "\n\n".join(paragraphs)
    payload.update(kind="document", text=text[:TEXT_PREVIEW_BYTES], truncated=len(text) > TEXT_PREVIEW_BYTES)
    return payload


def _pptx_payload(path: Path, payload: dict[str, Any]) -> dict[str, Any]:
    slides: list[str] = []
    with zipfile.ZipFile(path) as archive:
        names = sorted(
            (name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")),
            key=_slide_number,
        )
        for index, name in enumerate(names, 1):
            root = ElementTree.parse(archive.open(name)).getroot()
            text = " ".join(node.text or "" for node in root.iter() if node.tag.endswith("}t")).strip()
            slides.append(f"## Slide {index}\n\n{text}" if text else f"## Slide {index}")
            if sum(map(len, slides)) >= TEXT_PREVIEW_BYTES:
                break
    text = "\n\n".join(slides)
    payload.update(kind="document", text=text[:TEXT_PREVIEW_BYTES], truncated=len(text) > TEXT_PREVIEW_BYTES)
    return payload


def _slide_number(name: str) -> int:
    stem = Path(name).stem
    digits = "".join(character for character in stem if character.isdigit())
    return int(digits or 0)


def _archive_payload(path: Path, payload: dict[str, Any]) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    if zipfile.is_zipfile(path):
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            for info in infos[:ARCHIVE_ENTRY_LIMIT]:
                rows.append({"path": info.filename, "size": info.file_size, "compressed_size": info.compress_size})
            truncated = len(infos) > ARCHIVE_ENTRY_LIMIT
    else:
        with tarfile.open(path) as archive:
            members = archive.getmembers()
            for member in members[:ARCHIVE_ENTRY_LIMIT]:
                rows.append({"path": member.name, "size": member.size, "compressed_size": None})
            truncated = len(members) > ARCHIVE_ENTRY_LIMIT
    payload.update(kind="archive", columns=["path", "size", "compressed_size"], sample_rows=rows, truncated=truncated)
    if truncated:
        payload["warnings"].append(f"Archive preview shows at most {ARCHIVE_ENTRY_LIMIT} entries.")
    return payload
